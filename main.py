import asyncio
import os
import time
import openpyxl
import pandas as pd
import yaml
import logging

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from kinematics import rms
from kinemg_PDF import PDF_MM
from similarity import cmc
from yaml import SafeLoader

FORMAT = '%(asctime)s %(filename)3s %(user)-3s %(levelname)-3s: %(message)s'
logging.basicConfig(format=FORMAT)
d = {'user': os.getlogin()}
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

def find_folder():
    global path

    path = os.getcwd()      #os.path.dirname(__file__) --> another option
    for file in os.listdir(path):
        for strings in file.split(): 
            if strings.isdigit():
                folder = file
                return folder

            
def add_new_sheet(excel_book, df, sheetname, addnew = True):
    if addnew:
        excel_book.create_sheet(sheetname)
        ws = excel_book[sheetname]
    else: 
        ws = excel_book.active
        ws.title = sheetname

    rows = dataframe_to_rows(df)

    # For each row

    for r_idx, row in enumerate(rows, 1):

        # Write each cell for each column
        
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    if sheetname == 'Similitud':
        ws.column_dimensions['A'].width = 35
    else: 
        for col in ws.columns:
            column = col[0].column_letter
            if column != 'A':
                ws.column_dimensions[column].width = 20
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment =  Alignment(horizontal='center', wrap_text=True)
    return excel_book


async def save_excel(six, selected_run): 
    p_path = os.sep.join([path, folder])
    runs_mdx = os.listdir(p_path)[2:5]
    runs = [run.split()[-1][1:2] for run in runs_mdx]
    runs = [selected_run] + [run for run in runs if run != selected_run]          #runs with the selected run first

    kin_data = pd.read_fwf(f'{p_path}{os.sep}graficas{runs[0]}.emt', skiprows=6).drop(['Cycle'], axis=1).set_index(['Sample'])
    kin_data2 = pd.read_fwf(f'{p_path}{os.sep}graficas{runs[1]}.emt', skiprows=6).drop(['Cycle'], axis=1).set_index(['Sample'])
    kin_data3 = pd.read_fwf(f'{p_path}{os.sep}graficas{runs[2]}.emt', skiprows=6).drop(['Cycle'], axis=1).set_index(['Sample'])
    data_similarity = cmc(kin_data, kin_data2, kin_data3, runs)
    
    with open('config.yaml', encoding='utf8') as yaml_file:
        config = yaml.load(yaml_file, Loader=SafeLoader)
        kinematics_headers = config['Headers']['Kinematics']
        six_data_path = config['sixmin_folder_path']
    kin_data.columns = [kinematics_headers[joint] for joint in kin_data.columns]

    if six == 'SI':
        logger.info('Creating excel', extra=d)
        # six_data_path = r'C:\\Users\\marcha\\Desktop\\PACIENTES\\PACIENTES VAR_6_MINS'
        p_folder_name = ' '.join(folder.split()[:-1]) + ' 6_MIN'
        p_six_path = os.sep.join([six_data_path, p_folder_name, f'Data6_min_{file_name}.xlsx'])

        # Open six_min_data excel

        try: 
            excel_book = openpyxl.load_workbook(p_six_path)
        except FileNotFoundError: 
            p_folder_name = ' '.join(folder.split()[:-1]) + ' 6MIN'
            p_six_path = os.sep.join([six_data_path, p_folder_name, f'Data6_min_{file_name}.xlsx'])
            excel_book = openpyxl.load_workbook(p_six_path)

        # Create new sheets

        excel_book = add_new_sheet(excel_book, kin_data, 'Datos Cinemática') 
        excel_book = add_new_sheet(excel_book, data_similarity, 'Similitud') 

        excel_book.save(f'{p_path}{os.sep}Grafica&Data_6min_{file_name}.xlsx')
        logger.info('Excel created', extra=d)

    elif six == 'NO':
        logger.info('Creating excel', extra=d)
        wb = Workbook()
        wb = add_new_sheet(wb, kin_data, 'Datos Cinemática', addnew=False)
        wb = add_new_sheet(wb, data_similarity, 'Similitud')
        wb.save(f'{p_path}{os.sep}Graficas_{file_name}.xlsx')
        logger.info('Excel created', extra=d)
    else:
        logger.warning('Caution: Check in "NOTA.txt" 6min option, the excel file has not been created.', extra=d)
    

def p_data():
    global file_name, folder

    folder = find_folder()
    file_name = ' '.join(folder.split()[1:-1])

    
    p_id = [int(s) for s in folder.split() if s.isdigit()][0]

    p_name = file_name[:-1]
    try:
        with open(os.sep.join([path, folder, '/NOTA.txt']), encoding='utf8') as f:
            lines = f.read()
    except UnicodeDecodeError:
        with open(os.sep.join([path, folder, '/NOTA.txt']), encoding='latin-1') as f:
            lines = f.read()
    finally:
        age = [int(s) for s in lines.split() if s.isdigit()][0]
        for index, line in enumerate(lines.split()):
            if line == "DIAGNOSTICO:":
                index_0 = index + 1
            if line == "6MIN":
                index_f = index
            if line == "SENSOR:":
                index_6 = index + 1
            if line == 'INDIVIDUAL:':
                index_run = index + 1
        
        six_min = ' '.join(lines.split()[index_6:index_6 + 1])
        pathology = ' '.join(lines.split()[index_0:index_f])
        selected_run = ''.join(lines.split()[index_run:index_run + 1])
    asyncio.run(save_excel(six_min, selected_run))
    return p_name, age, pathology, p_id, selected_run


def main():
    t0 = time.time()
    if not os.path.exists('images'):
        os.mkdir('images')
    p_info = p_data()
    print(p_info)
    rms(p_info[1], os.sep.join([path, folder]), p_info[-1])
    # PDF
    pdf = PDF_MM(p_info, 'L', 'mm', 'letter')
    pdf.pag1()
    pdf.pag2()
    pdf.pag3()

    try:
        pdf.output(os.sep.join([path, folder, f'{file_name}_ReporteMM.pdf']))
    except PermissionError:
        logger.warning('You have an open pdf with the same name')

    logger.info(f'Documento creado en: {os.sep.join([path, folder])}', extra=d)
    logger.info(f'Tiempo transcurrido: {time.time() - t0}', extra=d)


if __name__ == "__main__":
    main()